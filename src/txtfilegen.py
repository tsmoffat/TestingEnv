"""A minor tool module for generating textfiles of combinations. Just for internal use really"""
import decimal as dec
import openpyxl as xl
from numbers import Number

workbook = xl.load_workbook("Five RF Sections - Relative Effects.xlsx")
s180 = workbook.get_sheet_by_name('180')
s90 = workbook.get_sheet_by_name('90')
s45 = workbook.get_sheet_by_name('45')
s225 = workbook.get_sheet_by_name('22.5')

f = open('18090.txt', 'w')
for i in s180.iter_rows(row_offset=2):
    for j in s90.iter_rows(row_offset=2):
        print(j)
        if i[2].value is not None and j[2].value is not None:
            convert1 = dec.Decimal(i[2].value)
            convert2 = dec.Decimal(j[2].value)
            total = convert1 + convert2
            total = total.quantize(dec.Decimal('.001'), rounding=dec.ROUND_HALF_UP)
            f.write(str(total) + "\n")





