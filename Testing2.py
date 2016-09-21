"""Module to work on more ideas for MPAC tuning."""
import os
import decimal as dec
import openpyxl as xl
from src import extras as ex
from src import TwoPhase as tp
from src import ThreePhase as thp
from src import FourPhase as fp
from src import AttenuationSearch as ats
from src import lookuptablegenerator as lutg
from src import minphasevariation as mpv
from src import mininsertloss as mil
from src import minphaseatt as mpa
from src import minattvar as mav
from src import numberformatting as nf


class Main:
    """The main controlling class."""

    def __init__(self):
        """Initialise class."""
        print("Initialising")
        self.phase24 = 1  # Sets up all the important values in trhe program, mostly referring to positions in the spreadsheet
        self.phase28 = 2
        self.phase32 = 3
        self.att24 = 4
        self.att28 = 5
        self.att32 = 6
        self.dsastate24 = 0
        self.dsas2124 = 2
        self.dsaphase2124 = 4
        self.dsastate28 = 6
        self.dsas2128 = 8
        self.dsaphase2128 = 10
        self.dsastate32 = 12
        self.dsas2132 = 14
        self.dsaphase2132 = 16
        self.targetphase = 0
        self.targetatt = 0
        self.workbook = xl.load_workbook(os.path.join(
            os.path.dirname(__file__), 'source.xlsx'))
        self.s180 = self.workbook.get_sheet_by_name('180')
        self.s90 = self.workbook.get_sheet_by_name('90')
        self.s45 = self.workbook.get_sheet_by_name('45')
        self.s225 = self.workbook.get_sheet_by_name('22.5')
        self.dsa = self.workbook.get_sheet_by_name('DSA')
        self.k = 0

    def main(self):
        """Control all the other modules in the program."""
        dec.getcontext().prec = 6
        consent = input(
            "Would you like to calculate a specific (V)alue or generate a lookup (T)able?")
        if consent == 'V':
            # Selects each module individually, there may be more added in in
            # the future
            print("Choose what you would like to find")
            print("E - the most efficient value")
            print("D - the value with least variation across frequency")
            print("I - the value with minimum insertion loss")
            print("A - the attenuation value with minimum variation across frequency")
            print(
                "P - the attenuation value with minimum phase difference across frequency")
            minvarchoice = input()

            list180 = []
            for row in self.s180.iter_rows(row_offset=2):
                if row[self.phase28].value is not None:
                    list180.append(dec.Decimal(row[self.phase28].value))
            set180 = set(list180)

            list90 = []
            for row in self.s90.iter_rows(row_offset=2):
                if row[self.phase28].value is not None:
                    list90.append(dec.Decimal(row[self.phase28].value))
            set90 = set(list90)

            list45 = []
            for row in self.s45.iter_rows(row_offset=2):
                if row[self.phase28].value is not None:
                    list45.append(dec.Decimal(row[self.phase28].value))
            set45 = set(list45)

            list225 = []
            for row in self.s225.iter_rows(row_offset=2):
                if row[self.phase28].value is not None:
                    list225.append(dec.Decimal(row[self.phase28].value))
            set225 = set(list225)
            if minvarchoice == "D":
                self.targetphase = input(
                    "Please enter the desired phase shift for 28GHz")
                self.k = input("How many values would you like to search?")
                minphase = mpv.minvariation(self, set180, set90, set45, set225)
                print("The closest value is " + str(minphase[
                      'total']) + ", giving a total variation of " + str(minphase['totalphasediff']))
                formatted = nf.phaseformat(self, minphase)
                print("The settings you need for this are:")
                print("180: " + formatted['s180setting'])
                print("90: " + formatted['s90setting'])
                print("45: " + formatted['s45setting'])
                print("22.5: " + formatted['s225setting'])
            elif minvarchoice == "E":
                self.targetphase = input(
                    "Please enter the desired phase shift for 28GHz")
                self.targetatt = input(
                    "Please enter the desired attenuation for 28GHz")
                bestresult = ex.checkall(self, set180, set90, set45, set225)
                bestresult2 = tp.checkall(self, set180, set90, set45, set225)
                bestresult3 = thp.checkall(self, set180, set90, set45, set225)
                closest = fp.closest_finder(self)
                bestresult4 = fp.check(
                    self, set180, set90, set45, set225, closest)
                sollist = [bestresult['total'], bestresult2['total'],
                           bestresult3['total'], bestresult4['total']]
                bestphase = ex.mostaccurate(
                    self, bestresult, bestresult2, bestresult3, bestresult4, sollist)
                attlist = ats.attlist(self)
                closest = ats.closest(self, attlist)
                bestatt = ats.attenuationsearch(self, attlist, closest)
                print("The best result is " + str(bestphase[
                      'total']) + " and the best attenuation is " + str(bestatt['att2128']))
                formatted = nf.phaseformat(self, bestphase)
                formatatt = nf.attformat(self, bestatt)
                print("The settings you need for this are: ")
                print("180: " + formatted['s180setting'])
                print("90: " + formatted['s90setting'])
                print("45: " + formatted['s45setting'])
                print("22.5: " + formatted['s225setting'])
                print("Attenuation: " + formatatt)
            elif minvarchoice == "I":
                self.targetphase = input(
                    "Please enter the desired phase shift for 28GHz")
                self.k = input("How many values would you like to search?")
                mininsert = mil.mininsertloss(
                    self, set180, set90, set45, set225)
                print("The best value is " + str(mininsert[
                      'total']) + ", giving a total insertion loss of " + str(mininsert['totalatt']))
                formatted = nf.phaseformat(self, mininsert)
                print("The settings you need for this are: ")
                print("180: " + formatted['s180setting'])
                print("90: " + formatted['s90setting'])
                print("45: " + formatted['s45setting'])
                print("22.5: " + formatted['s225setting'])
            elif minvarchoice == "P":
                self.targetatt = input(
                    "Please enter the desired attenuation for 28GHz")
                self.k = input("How many values would you like to search?")
                minamp = mpa.minampvar(self)
                print("The best value is " + str(minamp['att2128']) +
                      ", giving a variation of " + str(minamp['variation']))
                formatted = nf.attformat(self, minamp)
                print("The setting you need is: " + formatted)
            elif minvarchoice == "A":
                self.targetatt = input(
                    "Please enter the desired attenuation for 28GHz")
                self.k = input("How many values would you like to search?")
                minatt = mav.minattvar(self)
                print("The best attenuation is " +
                      str(minatt['att2128']) + ", giving a variation of " + str(minatt['variation']))
                resultssetting = nf.attformat(self, minatt)
                print("The setting you need is: " + resultssetting)

            else:
                print("Not a valid option")
        elif consent == 'T':
            option = input("Generate a table for (A)ttenuation or (P)hase?")
            if option == "A":
                lutg.atttablegen(self)
            elif option == "P":
                lutg.tablegen(self)
            else:
                print("That is not an option")

TESTBENCH = Main()
TESTBENCH.main()
