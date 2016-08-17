"""Module to work out the required settings for MPAC"""
from sys import exit
from decimal import Decimal, getcontext
import openpyxl as xl


class Main:
    """Where most of the actual work happens, this could definitely be split into subclasses"""

    def __init__(self):
        print("Initialising")

    def main(self):
        getcontext().prec = 6
        workbook = xl.load_workbook("Five RF Sections - Relative Effects.xlsx")
        s180 = workbook.get_sheet_by_name('180')
        s90 = workbook.get_sheet_by_name('90')
        s45 = workbook.get_sheet_by_name('45')
        s225 = workbook.get_sheet_by_name('22.5')

        frequency = input("What frequency are you using?")
        if frequency != "24GHz" and frequency != "28GHz" and frequency != "32GHz":
            frequency = input(
                'Sorry, that is not a valid frequency, please enter 24GHz, 28GHz or 32GHz')

        if frequency == '24GHz':
            phase = 1
            att = 5
            dsastate = 1
            dsas21 = 3

        elif frequency == '28GHz':
            phase = 2
            att = 6
            dsastate = 6
            dsas21 = 8
        elif frequency == '32GHz':
            phase = 3
            att = 7
            dsastate = 11
            dsas21 = 13
        else:
            phase = 0
            att = 0
            dsastate = 0
            dsas21 = 0

        list180 = []
        for row in s180.iter_rows(row_offset=2):
            if row[phase].value is not None:
                list180.append(Decimal(row[phase].value))
        set180 = set(list180)

        list90 = []
        for row in s90.iter_rows(row_offset=2):
            if row[phase].value is not None:
                list90.append(Decimal(row[phase].value))
        set90 = set(list90)

        list45 = []
        for row in s45.iter_rows(row_offset=2):
            if row[phase].value is not None:
                list45.append(Decimal(row[phase].value))
        set45 = set(list45)

        list225 = []
        for row in s225.iter_rows(row_offset=2):
            if row[phase].value is not None:
                list225.append(Decimal(row[phase].value))
        set225 = set(list225)
        constants = {'s180': s180, 's90': s90, 's45': s45}
        constants['s225'] = s225
        constants['set180'] = set180
        constants['set90'] = set90
        constants['set45'] = set45
        constants['set225'] = set225
        constants['phase'] = phase
        constants['att'] = att

        targetphase = Decimal(input("What phase do you want to achieve?"))
        targetatt = input("What attenuation do you want to achieve?")
        constants['targetphase'] = targetphase
        constants['targetatt'] = targetatt
        result18 = check180(list180, s180,)

        if targetphase in set90:
            print("Found it!")
            for row in s90.iter_rows():
                if row[phase].value == targetphase:
                    optimalrow = int(row[0].value)
                    att90 = Decimal(row[att].value)
            print("Optimal solution found, it is state " + str(optimalrow))
            print(att90)
            exit()

        elif targetphase in set45:
            print("Found it!")
            for row in s45.iter_rows():
                if row[phase].value == targetphase:
                    optimalrow = int(row[0].value)
                    att45 = Decimal(row[att].value)
            print("Optimal solution found, it is state " + str(optimalrow))
            print(att45)
            exit()

        elif targetphase in set225:
            print("Found it!")
            for row in s225.iter_rows():
                if row[phase].value == targetphase:
                    optimalrow = int(row[0].value)
                    att225 = Decimal(row[att].value)
            print("Optimal solution found, it is state " + str(optimalrow))
            print(att225)
            exit()

        else:

            sum2 = two_element_sum(constants)
            resultsdict = sum2.Two_element_sum()
            results1 = resultsdict['solution1']
            if results1:
                print(results1)
                results2 = resultsdict['solution2']
                row1 = resultsdict['row1']
                row2 = resultsdict['row2']
                att1 = resultsdict['att1']
                att2 = resultsdict['att2']
                print(row1, row2, results2, att1, att2)
                exit()
            else:
                print("This is a placeholder for now")
                exit()

    def check180(self, list180, s180, targetphase, targetatt, phase, att):
        if targetphase in list180:
            print("Found it!")
            for row in s180.iter_rows():
                if row[phase].value == targetphase:
                    row180 = int(row[0].value)
                    att180 = Decimal(row[att].value)

            print("Optimal solution found, it is state " + str(row180))
            print(att180)
            return {'row180': row180, 'att180': att180}
            exit()
        else:
            closest = min(list180, key=lambda x: abs(x - targetphase))
            for row in self.s180.iter_rows:
                if row[phase].value == closest:
                    row1 = int(row[0].value)
                    att1 = Decimal(row[att].value)




class two_element_sum(Main):
    """For working out the two element sums"""

    def __init__(self, constants):
        getcontext().prec = 6
        self.s180 = constants['s180']
        self.s90 = constants['s90']
        self.s45 = constants['s45']
        self.s225 = constants['s225']
        self.set180 = constants['set180']
        self.set90 = constants['set90']
        self.set45 = constants['set45']
        self.set225 = constants['set225']
        self.targetphase = constants['targetphase']
        self.phase = constants['phase']
        self.att = constants['att']

    def Two_element_sum(self):
        getcontext().prec = 6
        solution1 = 0
        for i in self.set180:
            for j in self.set90:

                result = i + j

                if result == self.targetphase:
                    solution1 = i
                    solution2 = j
                    print("Found it! It requires two vectors")
                    print("180 and 90")
                    for row in self.s180.iter_rows():
                        if row[self.phase].value == i:
                            row1 = int(row[0].value)
                            att1 = Decimal(row[self.att].value)

                    for row in self.s90.iter_rows():
                        if row[self.phase].value == j:
                            row2 = int(row[0].value)
                            att2 = Decimal(row[self.att].value)

        for i in self.set180:
            for j in self.set45:
                result = i + j
                if result == self.targetphase:
                    solution1 = i
                    solution2 = j
                    print("Found it! It requires two vectors")
                    print("180 and 45")
                    for row in self.s180.iter_rows():
                        if row[self.phase].value == i:
                            row1 = int(row[0].value)
                            att1 = Decimal(row[self.att].value)
                            print(row1)
                            print(att1)

                    for row in self.s45.iter_rows():
                        if row[self.phase].value == j:
                            row2 = int(row[0].value)
                            att2 = Decimal(row[self.att].value)
                            print(row2)
                            print(att2)

        for i in self.set180:
            for j in self.set225:
                result = i + j
                if result == self.targetphase:
                    solution1 = i
                    solution2 = j
                    print("Found it! It requires two vectors")
                    print("180 and 22.5")
                    for row in self.s180.iter_rows():
                        if row[self.phase].value == i:
                            row1 = int(row[0].value)
                            att1 = Decimal(row[self.att].value)
                            print(row1)
                            print(att1)

                    for row in self.s225.iter_rows():
                        if row[self.phase].value == i:
                            row2 = int(row[0].value)
                            att2 = Decimal(row[self.att].value)
                            print(row2)
                            print(att2)

        for i in self.set90:
            for j in self.set45:
                result = i + j
                if result == self.targetphase:
                    solution1 = i
                    solution2 = j
                    print("Found it! It requires two vectors")
                    print("90 and 45")
                    for row in self.s90.iter_rows():
                        if row[self.phase].value == i:
                            row1 = int(row[0].value)
                            att1 = Decimal(row[self.att].value)
                            print(row1)
                            print(att1)

                    for row in self.s45.iter_rows():
                        if row[self.phase].value == j:
                            row2 = int(row[0].value)
                            att2 = Decimal(row[self.att].value)
                            print(row2)
                            print(att2)
        for i in self.set90:
            for j in self.set225:
                result = i + j
                if result == self.targetphase:
                    solution1 = i
                    solution2 = j
                    print("Found it! It requires two vectors")
                    print("90 and 22.5")
                    for row in self.s90.iter_rows():
                        if row[self.phase].value == i:
                            row1 = int(row[0].value)
                            att1 = Decimal(row[self.att].value)
                            print(row1)
                            print(att1)

                    for row in self.s225.iter_rows():
                        if row[self.phase].value == j:
                            row2 = int(row[0].value)
                            att2 = Decimal(row[self.att].value)
                            print(row2)
                            print(att2)

        for i in self.set45:
            for j in self.set225:
                result = i + j
                if result == self.targetphase:
                    solution1 = i
                    solution2 = j
                    print("Found it! It requires two vectors")
                    print("45 and 22.5")
                    for row in self.s90.iter_rows():
                        if row[self.phase].value == i:
                            row1 = int(row[0].value)
                            att1 = Decimal(row[self.att].value)

                    for row in self.s90.iter_rows():
                        if row[self.phase].value == i:
                            row2 = int(row[0].value)
                            att2 = Decimal(row[self.att].value)

        if solution1:
            return {'solution1': solution1, 'solution2': solution2, 'row1': row1, 'att1': att1, 'row2': row2, 'att2': att2}


class three_element_sum(Main):
    def __init__(self, s180, s90, s45, s225, set180, set90, set45, set225, targetphase, phase, att):
        getcontext().prec = 6
        self.s180 = s180
        self.s90 = s90
        self.s45 = s45
        self.s225 = s225
        self.set180 = set180
        self.set90 = set90
        self.set45 = set45
        self.set225 = set225
        self.targetphase = targetphase
        self.phase = phase
        self.att = att

    def three_element_sum(self):
        getcontext().prec = 6
        solution1 = 0
        for i in self.set180:
            for j in self.set90:
                for k in self.set45:
                    result = i + j + k
                    if result == self.targetphase:
                        solution1 = i
                        solution2 = j
                        solution3 = k
                        print("Found it! It requires three vectors")
                        print("180, 90 and 45")

                        for row in self.s180.iter_rows():
                            if row[self.phase].values == i:
                                row1 = int(row[0].value)
                                att1 = Decimal(row[self.att].value)

                        for row in self.s90.iter_rows():
                            if row[self.phase].values == j:
                                row2 = int(row[0].value)
                                att2 = Decimal(row[self.att].value)

                        for row in self.s45.iter_rows():
                            if row[self.phase].values == k:
                                row3 = int(row[0].value)
                                att3 = Decimal(row[self.att].value)
        for i in self.set180:
            for j in self.set45:
                for k in self.set225:
                    result = i + j + k
                    if result == self.targetphase:
                        solution1 = i
                        solution2 = j
                        solution3 = k
                        print("Found it! It requires three vectors")
                        print("180, 45 and 22.5")

                        for row in self.s180.iter_rows():
                            if row[self.phase].values == i:
                                row1 = int(row[0].value)
                                att1 = Decimal(row[self.att].value)

                        for row in self.s45.iter_rows():
                            if row[self.phase].values == j:
                                row2 = int(row[0].value)
                                att2 = Decimal(row[self.att].value)

                        for row in self.s225.iter_rows():
                            if row[self.phase].values == k:
                                row3 = int(row[0].value)
                                att3 = Decimal(row[self.att].value)


TESTBENCH = Main()
TESTBENCH.main()
